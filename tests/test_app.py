from datetime import date, datetime, timezone
from io import BytesIO
from decimal import Decimal
from pathlib import Path
import re

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pydantic import ValidationError
from starlette.requests import Request

from app import main as main_module
from app.main import (
    METER_KEYS,
    METERS,
    ReadingPayload,
    build_workbook,
    export_month_options,
    fetch_month,
    month_bounds,
    resolve_public_url,
    validate_settings,
)


def sample_record(day: int, factor: int = 1) -> dict:
    return {
        "date": f"2026-08-{day:02d}",
        "time": "08:00",
        "reader": "测试人员",
        "values": {key: Decimal(index * factor) for index, key in enumerate(METER_KEYS, 1)},
    }


def test_payload_requires_all_nonnegative_meter_values():
    payload = ReadingPayload(reader=" 测试人员 ", values=sample_record(1)["values"])
    assert payload.reader == "测试人员"
    assert len(payload.values) == 18
    with pytest.raises(ValidationError):
        ReadingPayload(reader="测试人员", values={"boka": Decimal("1")})
    invalid = sample_record(1)["values"]
    invalid["boka"] = Decimal("-1")
    with pytest.raises(ValidationError):
        ReadingPayload(reader="测试人员", values=invalid)


def test_settings_require_distinct_secrets(monkeypatch):
    monkeypatch.setenv("DATABASE_URL", "postgresql://example")
    monkeypatch.setenv("SUBMISSION_CODE", "same-secret")
    monkeypatch.setenv("ADMIN_PASSWORD", "same-secret")
    with pytest.raises(RuntimeError, match="must be different"):
        validate_settings()


def test_frontend_contains_every_backend_meter():
    html = (Path(__file__).parents[1] / "public" / "index.html").read_text(encoding="utf-8")
    frontend_keys = tuple(re.findall(r'\{ id:"([^"]+)", group:', html))
    assert frontend_keys == METER_KEYS
    assert '<select id="exportMonth" aria-label="选择月份">' in html
    assert ".month-control { position:relative; display:block; width:100%; height:46px;" in html
    assert "text-align:center; text-align-last:center;" in html
    assert "populateMonthOptions(body.exportMonths);" in html


def test_meter_multipliers_match_business_rules():
    multipliers = {key: multiplier for key, _, multiplier in METERS}

    assert multipliers["boka"] == multipliers["huaman"] == 3000
    assert {multipliers[f"solar{index}"] for index in range(1, 5)} == {1}
    assert {multipliers[f"charger{index}"] for index in range(1, 4)} == {1}


def test_fetch_month_serializes_a_populated_result(monkeypatch):
    row = {
        "reading_date": date(2026, 8, 29),
        "recorded_at": datetime(2026, 8, 29, 8, 38, tzinfo=timezone.utc),
        "reader": "刘婷婷",
        **{key: Decimal(index) for index, key in enumerate(METER_KEYS, 1)},
    }

    class Result:
        def fetchall(self):
            return [row]

    class Connection:
        def __enter__(self):
            return self

        def __exit__(self, *_):
            return None

        def execute(self, *_):
            return Result()

    monkeypatch.setattr("app.main.connect", lambda: Connection())

    records = fetch_month("2026-08")

    assert records == [{
        "date": "2026-08-29",
        "time": "16:38",
        "reader": "刘婷婷",
        "values": {key: Decimal(index) for index, key in enumerate(METER_KEYS, 1)},
    }]


def test_month_bounds_handles_december():
    assert month_bounds("2026-08") == ("2026-08-01", "2026-09-01")
    assert month_bounds("2026-12") == ("2026-12-01", "2027-01-01")


def test_export_months_include_history_and_exclude_future():
    months = export_month_options(datetime(2026, 8, 31, tzinfo=timezone.utc))

    assert months[:2] == ["2026-08", "2026-07"]
    assert "2026-01" in months
    assert "2025-12" in months
    assert "2026-09" not in months
    assert months[-1] == "1900-01"


def test_export_matches_template_and_keeps_formula_results():
    content = build_workbook([sample_record(1), sample_record(2, 2)])
    formulas = load_workbook(BytesIO(content), data_only=False)["每日抄表"]
    values = load_workbook(BytesIO(content), data_only=True)["每日抄表"]

    assert formulas.max_column == 20
    assert formulas["A1"].value == "每日水电用量抄表记录"
    assert formulas["D3"].value == "一号并网柜"
    assert formulas["B4"].value == 3000
    assert formulas["D4"].value == 3
    assert formulas["H4"].value == 7
    assert formulas["K4"].value == 10
    assert formulas["B26"].value == "=SUM(B4:B25)"
    assert formulas["D26"].value == "=SUM(D4:D25)"
    assert formulas["K26"].value == "=SUM(K4:K25)"
    assert values["B26"].value == 9000
    assert values["D26"].value == 9
    assert values["K26"].value == 30
    assert formulas["A1"].font.name == "Noto IKEA Simplified Chinese"
    assert {str(item) for item in formulas.merged_cells.ranges} == {
        "A1:S1", "A2:A3", "B2:C2", "D2:G2", "H2:J2", "K2:P2", "Q2:S2", "T2:T3"
    }


def test_public_url_uses_current_public_https_host(monkeypatch):
    monkeypatch.setenv("PUBLIC_URL", "http://old-internal-host:8080")
    request = Request({
        "type": "http",
        "scheme": "http",
        "path": "/api/qr.png",
        "query_string": b"",
        "headers": [(b"host", b"meter.example.com")],
        "server": ("internal-service", 8080),
    })

    assert resolve_public_url(request) == "https://meter.example.com"


def test_public_url_honors_forwarded_host():
    request = Request({
        "type": "http",
        "scheme": "http",
        "path": "/api/qr.png",
        "query_string": b"",
        "headers": [
            (b"host", b"internal-service:8080"),
            (b"x-forwarded-host", b"meter.example.com"),
        ],
        "server": ("internal-service", 8080),
    })

    assert resolve_public_url(request) == "https://meter.example.com"


def test_native_form_export_returns_mobile_friendly_attachment(monkeypatch):
    monkeypatch.setenv("ADMIN_PASSWORD", "admin-secret")
    monkeypatch.setattr(main_module, "fetch_month", lambda _: [sample_record(1)])

    response = TestClient(main_module.app).post(
        "/api/export",
        data={"month": "2026-08", "admin_password": "admin-secret"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'filename="meter-reading-2026-08.xlsx"' in response.headers["content-disposition"]
    assert "filename*=UTF-8''" in response.headers["content-disposition"]
    assert int(response.headers["content-length"]) == len(response.content)
    assert load_workbook(BytesIO(response.content), data_only=True)["每日抄表"]["B4"].value == 3000


def test_signed_get_export_returns_android_download(monkeypatch):
    monkeypatch.setenv("ADMIN_PASSWORD", "admin-secret")
    monkeypatch.setattr(main_module, "fetch_month", lambda _: [sample_record(1)])
    client = TestClient(main_module.app)

    ticket = client.post(
        "/api/export-ticket?month=2026-08",
        headers={"X-Admin-Password": "admin-secret"},
    )
    response = client.get(ticket.json()["downloadUrl"])

    assert ticket.status_code == 200
    assert response.status_code == 200
    assert response.headers["content-transfer-encoding"] == "binary"
    assert 'filename="meter-reading-2026-08.xlsx"' in response.headers["content-disposition"]
    assert load_workbook(BytesIO(response.content), data_only=True)["每日抄表"]["B4"].value == 3000


def test_signed_get_export_rejects_tampered_or_expired_ticket(monkeypatch):
    monkeypatch.setenv("ADMIN_PASSWORD", "admin-secret")
    client = TestClient(main_module.app)
    expired = int(main_module.time.time()) - 1
    signature = main_module.sign_download_ticket("2026-08", expired)

    expired_response = client.get(
        f"/api/export-download?month=2026-08&expires={expired}&signature={signature}"
    )
    tampered_response = client.get(
        f"/api/export-download?month=2026-08&expires={expired + 60}&signature={signature}"
    )

    assert expired_response.status_code == 401
    assert tampered_response.status_code == 401


def test_qr_print_page_instructs_xiaomi_users_to_use_system_camera():
    response = TestClient(main_module.app).get("/qr")

    assert response.status_code == 200
    assert '<img src="/api/qr.png"' in response.text
    assert "请使用手机系统相机扫码" in response.text
    assert "请勿使用微信" in response.text
